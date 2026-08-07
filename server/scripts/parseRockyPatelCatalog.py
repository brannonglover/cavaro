#!/usr/bin/env python3
"""Parse Rocky Patel blend sheets xlsx into catalog.json + processed product images."""

from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from PIL import Image

XLSX = Path("/Users/bglover/Downloads/RP Blend Sheets 2026.xlsx")
ROOT = Path(__file__).resolve().parents[2] / "tmp" / "rocky-patel"
RAW = ROOT / "raw-images"
PROC = ROOT / "processed"
ASSETS = Path(__file__).resolve().parents[2] / "assets" / "rocky-patel"

FRAC = {
    "½": "1/2",
    "¼": "1/4",
    "¾": "3/4",
    "⅛": "1/8",
    "⅜": "3/8",
    "⅝": "5/8",
    "⅞": "7/8",
}
HEADER_WORDS = {
    "BRAND",
    "SIZE",
    "DIMENSION",
    "ORIGIN",
    "TALKING POINTS",
    "WRAPPER | BINDER | FILLER",
    "WRAPPER|BINDER|FILLER",
}
TITLE_FIXES = {
    "DARK STAR": "Dark Star",
    "CONVICTION": "Conviction",
    "NUMBER 6": "Number 6",
    "A.L.R (Aged, Limited & Rare) 2ND EDITION": "A.L.R. 2nd Edition",
    "Edición Unica": "Edicion Unica",
    "25th Year by Hamlet Pardes": "25th Year by Hamlet Paredes",
    "Vintage 2006 San Andreas": "Vintage 2006 San Andres",
}
SHEET_PRIORITY = {
    "Anniversary Series (2)": 2,
}
BG = np.array([0x21, 0x19, 0x12], dtype=np.float32)


def normalize_fraction_text(s: str) -> str:
    s = str(s).strip()
    for k, v in FRAC.items():
        s = s.replace(k, v)
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s*[xX×]\s*", "x", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_length(dim: str) -> str | None:
    if not dim:
        return None
    s = normalize_fraction_text(dim)
    m = re.match(
        r"^(\d+(?:\s+\d+/\d+)?)x(\d+(?:/\d+)?(?:-\d+)?)$",
        s,
        re.I,
    )
    if not m:
        return None
    return f"{m.group(1)}x{m.group(2)}"


def slugify(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def clean_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def extract_components(cell_text: str):
    wrapper = binder = filler = ""
    if not cell_text:
        return wrapper, binder, filler
    t = clean_text(cell_text)
    for part in re.split(r"(?=WRAPPER:|BINDER:|FILLER:)", t, flags=re.I):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"(WRAPPER|BINDER|FILLER)\s*:\s*(.+)$", part, re.I)
        if not m:
            continue
        kind, val = m.group(1).upper(), clean_text(m.group(2))
        if kind == "WRAPPER" and not wrapper:
            wrapper = val
        elif kind == "BINDER" and not binder:
            binder = val
        elif kind == "FILLER" and not filler:
            filler = val
    return wrapper, binder, filler


def infer_name_from_talking(talk: str) -> str | None:
    t = clean_text(talk)
    if not t:
        return None
    m = re.search(
        r"[“\"]Year of the ([^”\"]+)[”\"]|Year of the ([A-Za-z]+)",
        t,
    )
    if m:
        animal = next(g for g in m.groups() if g)
        return f"Year of the {animal.strip()}"
    if "Fifty was produced" in t:
        return "Fifty"

    # "Rocky Patel Sapphire is..." / "The Rocky Patel Emerald stands..."
    m = re.match(
        r"^(?:The\s+)?Rocky Patel\s+([A-Z][A-Za-z0-9’']+(?:\s+[A-Z][A-Za-z0-9’']+){0,4})\b",
        t,
    )
    if m:
        return m.group(1).strip()

    # "Gold Label delivers..." / "Dark Star captivates..."
    m = re.match(
        r"^([A-Z][A-Za-z0-9’'&.-]+(?:\s+[A-Z0-9][A-Za-z0-9’'&.-]*){0,5})\s+"
        r"(?:delivers|is|stands|showcases|features|offers)\b",
        t,
    )
    if m:
        return m.group(1).strip()

    return None


def display_name(raw: str) -> str:
    name = TITLE_FIXES.get(raw, raw)
    if name.isupper() and len(name) > 3:
        name = name.title()
    return name


def derive_line(name: str) -> str:
    if name.startswith("The Edge "):
        return "The Edge"
    if name.startswith("Vintage "):
        return "Vintage Series"
    if name.startswith("Java "):
        return "Java"
    if name.startswith("Catch 22"):
        return "Catch 22"
    if "Year of the" in name:
        return "Year of the"
    if name in ("It's A Boy", "It's A Girl"):
        return "It's A Boy/Girl"
    return name


def sheet_brand_col(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=12, max_col=7):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().upper() == "BRAND":
                return cell.column
    return 1


def anchor_row_col(img):
    a = img.anchor
    if isinstance(a, (TwoCellAnchor, OneCellAnchor)):
        return a._from.row + 1, a._from.col + 1
    return None, None


def recolor_and_crop(src_path: Path, dest_path: Path):
    arr = np.asarray(Image.open(src_path).convert("RGB"), dtype=np.float32)
    r, g, b = arr[..., 0], arr[..., 1], arr[..., 2]
    mx = np.maximum(np.maximum(r, g), b)
    mn = np.minimum(np.minimum(r, g), b)
    chroma = mx - mn
    luma = (r + g + b) / 3.0

    low, high = 198.0, 246.0
    t = np.clip((luma - low) / (high - low), 0.0, 1.0)
    t = t * t * (3.0 - 2.0 * t)
    chroma_keep = np.clip((chroma - 18.0) / 40.0, 0.0, 1.0)
    t = t * (1.0 - chroma_keep)

    mask = (t > 0.15).astype(np.uint8)
    padded = np.pad(mask, 1, mode="edge")
    dilated = np.maximum.reduce(
        [
            padded[0:-2, 0:-2],
            padded[0:-2, 1:-1],
            padded[0:-2, 2:],
            padded[1:-1, 0:-2],
            padded[1:-1, 1:-1],
            padded[1:-1, 2:],
            padded[2:, 0:-2],
            padded[2:, 1:-1],
            padded[2:, 2:],
        ]
    ).astype(np.float32)
    fringe = np.clip(dilated - mask.astype(np.float32), 0.0, 1.0)
    fringe_t = fringe * np.clip((luma - 170.0) / 60.0, 0.0, 1.0) * 0.85
    t = np.maximum(t, fringe_t)
    arr = arr * (1.0 - t[..., None]) + BG * t[..., None]

    dist = np.linalg.norm(arr - BG, axis=2)
    bbox = None
    for thr in (22, 16, 12, 9, 6):
        subject = dist > thr
        h, w = subject.shape
        border = max(1, min(h, w) // 80)
        subject[:border, :] = False
        subject[-border:, :] = False
        subject[:, :border] = False
        subject[:, -border:] = False
        ys, xs = np.where(subject)
        if len(xs) < 200:
            continue
        x0, x1 = int(xs.min()), int(xs.max())
        y0, y1 = int(ys.min()), int(ys.max())
        if (x1 - x0) < w * 0.015 or (y1 - y0) < h * 0.04:
            continue
        bbox = (x0, y0, x1, y1)
        break

    if bbox is not None:
        x0, y0, x1, y1 = bbox
        bw, bh = x1 - x0 + 1, y1 - y0 + 1
        h, w = arr.shape[:2]
        cigar_width_frac = 0.55
        vert_pad_frac = 0.06
        pad_y = max(8, int(bh * vert_pad_frac))
        target_w = max(bw + 16, int(bw / cigar_width_frac))
        pad_x = max(8, (target_w - bw) // 2)
        x0 = max(0, x0 - pad_x)
        y0 = max(0, y0 - pad_y)
        x1 = min(w - 1, x1 + pad_x)
        y1 = min(h - 1, y1 + pad_y)
        crop = arr[y0 : y1 + 1, x0 : x1 + 1]
        ch, cw = crop.shape[:2]
        want_w = max(cw, int(bw / cigar_width_frac))
        canvas = np.zeros((ch, want_w, 3), dtype=np.float32)
        canvas[:] = BG
        ox = (want_w - cw) // 2
        canvas[0:ch, ox : ox + cw] = crop
        arr = canvas

    Image.fromarray(np.clip(arr, 0, 255).astype(np.uint8), "RGB").save(
        dest_path, "JPEG", quality=90, optimize=True
    )


def score_image(rec) -> int:
    score = 0
    if 500 <= rec["w"] <= 800 and 250 <= rec["h"] <= 450:
        score += 5
    if rec["h"] > rec["w"]:
        score += 2
    if rec["h"] < 150:
        score -= 10
    return score


def parse_sheet(ws, sheet_name: str):
    brand_col = sheet_brand_col(ws)
    size_col = brand_col + 1
    dim_col = brand_col + 2
    wbf_col = brand_col + 4
    talk_col = brand_col + 5

    products = []
    line_meta = {}

    current = {
        "name": None,
        "start": None,
        "wrapper": "",
        "binder": "",
        "filler": "",
        "desc": "",
        "rows": [],
    }

    def flush():
        if not current["name"] or not current["rows"]:
            current["name"] = None
            current["rows"] = []
            return
        key = current["name"]
        end_row = current["rows"][-1]["row"]
        if key not in line_meta:
            line_meta[key] = {
                "name": key,
                "wrapper": current["wrapper"],
                "binder": current["binder"],
                "filler": current["filler"],
                "description": current["desc"],
                "sheet": sheet_name,
                "start_row": current["start"],
                "end_row": end_row,
            }
        else:
            meta = line_meta[key]
            if meta["sheet"] == sheet_name:
                meta["end_row"] = max(meta["end_row"], end_row)
            for field, value in (
                ("wrapper", current["wrapper"]),
                ("binder", current["binder"]),
                ("filler", current["filler"]),
                ("description", current["desc"]),
            ):
                if not meta.get(field) and value:
                    meta[field] = value
        products.extend(current["rows"])
        current["name"] = None
        current["rows"] = []

    def start_line(name, row, wrapper, binder, filler, desc):
        flush()
        current["name"] = name
        current["start"] = row
        current["wrapper"] = wrapper
        current["binder"] = binder
        current["filler"] = filler
        current["desc"] = desc

    def absorb_components(wrapper, binder, filler):
        if wrapper:
            current["wrapper"] = current["wrapper"] or wrapper
        if binder:
            current["binder"] = current["binder"] or binder
        if filler:
            current["filler"] = current["filler"] or filler

    for r in range(1, (ws.max_row or 0) + 1):
        brand_val = clean_text(ws.cell(r, brand_col).value)
        size_val = clean_text(ws.cell(r, size_col).value)
        dim_val = clean_text(ws.cell(r, dim_col).value)
        wbf_val = clean_text(ws.cell(r, wbf_col).value)
        talk_val = clean_text(ws.cell(r, talk_col).value)

        if brand_val.upper() in HEADER_WORDS or size_val.upper() in HEADER_WORDS:
            continue

        wrapper, binder, filler = extract_components(wbf_val)
        absorb_components(wrapper, binder, filler)

        new_brand = None
        if brand_val and brand_val.upper() not in HEADER_WORDS:
            new_brand = brand_val

        if not new_brand and talk_val and (size_val or dim_val):
            inferred = infer_name_from_talking(talk_val)
            if inferred and current["name"] != inferred:
                new_brand = inferred

        if new_brand:
            start_line(new_brand, r, wrapper, binder, filler, talk_val)

        if current["name"] is None and talk_val:
            inferred = infer_name_from_talking(talk_val)
            if inferred:
                start_line(inferred, r, wrapper, binder, filler, talk_val)

        if current["name"] is None and size_val and dim_val and talk_val:
            inferred = infer_name_from_talking(talk_val)
            if inferred:
                start_line(inferred, r, wrapper, binder, filler, talk_val)

        if not current["name"]:
            continue

        if talk_val and not current["desc"]:
            current["desc"] = talk_val
        absorb_components(wrapper, binder, filler)

        length = parse_length(dim_val) if dim_val else None
        if size_val and length:
            current["rows"].append(
                {
                    "brand": "Rocky Patel",
                    "name": current["name"],
                    "line": current["name"],
                    "size_name": size_val,
                    "length": length,
                    "sheet": sheet_name,
                    "row": r,
                }
            )
        elif size_val and dim_val and not length:
            print(
                f"WARN unparsed dim sheet={sheet_name} row={r} "
                f"size={size_val!r} dim={dim_val!r}"
            )

    flush()
    return products, line_meta


def main():
    RAW.mkdir(parents=True, exist_ok=True)
    PROC.mkdir(parents=True, exist_ok=True)
    ASSETS.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX)

    image_records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        brand_col = sheet_brand_col(ws)
        for j, img in enumerate(ws._images):
            row, col = anchor_row_col(img)
            data = img._data()
            im = Image.open(io.BytesIO(data))
            w, h = im.size
            near_brand = col is not None and col <= max(2, brand_col)
            aspect = w / max(h, 1)
            is_banner = aspect > 5.0 or h < 120
            if near_brand and not is_banner:
                kind = "product"
            elif near_brand and h > w and h >= 300:
                kind = "product"
            elif is_banner and near_brand:
                kind = "skip"
            else:
                kind = "blend"
            fname = f"{slugify(sheet_name)}_{j:02d}_r{row}_c{col}_{kind}.jpg"
            path = RAW / fname
            im.convert("RGB").save(path, "JPEG", quality=95)
            image_records.append(
                {
                    "sheet": sheet_name,
                    "idx": j,
                    "row": row,
                    "col": col,
                    "kind": kind,
                    "file": fname,
                    "path": str(path),
                    "w": w,
                    "h": h,
                }
            )

    print(
        "Images:",
        {
            k: sum(1 for r in image_records if r["kind"] == k)
            for k in ("product", "blend", "skip")
        },
    )

    all_products = []
    line_meta = {}
    for sheet_name in wb.sheetnames:
        products, meta = parse_sheet(wb[sheet_name], sheet_name)
        all_products.extend(products)
        for key, value in meta.items():
            if key not in line_meta:
                line_meta[key] = value
            else:
                existing = line_meta[key]
                if SHEET_PRIORITY.get(value["sheet"], 0) < SHEET_PRIORITY.get(
                    existing["sheet"], 0
                ):
                    line_meta[key] = value
                for field in ("wrapper", "binder", "filler", "description"):
                    if not existing.get(field) and value.get(field):
                        existing[field] = value[field]

    boy_rows = [p for p in all_products if p["name"] == "It's A Boy"]
    girl_rows = [p for p in all_products if p["name"] == "It's A Girl"]
    if boy_rows and not girl_rows:
        for br in boy_rows:
            all_products.append({**br, "name": "It's A Girl", "line": "It's A Girl"})
        line_meta["It's A Girl"] = {
            **line_meta.get("It's A Boy", {}),
            "name": "It's A Girl",
            "description": line_meta.get("It's A Boy", {}).get("description", ""),
        }

    catalog_map = {}
    for p in all_products:
        raw_name = p["name"]
        blend = display_name(raw_name)
        size_name = p["size_name"]
        # Blend name only; vitola lives in size_name so Add Cigar can list all sizes.
        name = blend
        meta = line_meta.get(raw_name) or {}
        for k, v in line_meta.items():
            if display_name(k) == blend or k == raw_name:
                meta = v
                break

        wrapper = meta.get("wrapper") or ""
        binder = meta.get("binder") or ""
        filler = meta.get("filler") or ""
        desc = meta.get("description") or ""
        if not wrapper:
            for token in (
                "Maduro",
                "Connecticut",
                "Cameroon",
                "Sumatra",
                "Corojo",
                "Habano",
                "Candela",
                "San Andrés",
                "San Andres",
                "Broadleaf",
            ):
                if token.lower() in blend.lower():
                    wrapper = token.replace("Andres", "Andrés")
                    break

        key = (name.lower(), size_name.lower(), p["length"])
        candidate = {
            "brand": "Rocky Patel",
            "name": name,
            "line": derive_line(blend),
            "description": desc,
            "wrapper": wrapper,
            "binder": binder,
            "filler": filler,
            "length": p["length"],
            "size_name": size_name,
            "image": "",
            "_sheet": p["sheet"],
            "_row": p["row"],
            "_size_name": size_name,
            "_raw_name": raw_name,
            "_blend": blend,
        }
        prev = catalog_map.get(key)
        if not prev:
            catalog_map[key] = candidate
            continue
        prev_pri = SHEET_PRIORITY.get(prev["_sheet"], 0)
        new_pri = SHEET_PRIORITY.get(candidate["_sheet"], 0)
        if new_pri < prev_pri or (
            new_pri == prev_pri
            and len(candidate["description"]) > len(prev["description"])
        ):
            catalog_map[key] = candidate
        else:
            for field in ("wrapper", "binder", "filler", "description"):
                if not prev[field] and candidate[field]:
                    prev[field] = candidate[field]

    catalog = list(catalog_map.values())

    starts_by_sheet = defaultdict(list)
    for raw, meta in line_meta.items():
        starts_by_sheet[meta["sheet"]].append(
            {
                "name": display_name(raw),
                "raw": raw,
                "start": meta["start_row"],
                "end": meta["end_row"],
            }
        )
    for sheet in starts_by_sheet:
        starts_by_sheet[sheet].sort(key=lambda x: x["start"])

    # Ensure Year-of / inferred lines appear in starts for image matching
    for p in catalog:
        sheet = p["_sheet"]
        blend = p["_blend"]
        if any(b["name"] == blend for b in starts_by_sheet[sheet]):
            continue
        starts_by_sheet[sheet].append(
            {
                "name": blend,
                "raw": p["_raw_name"],
                "start": p["_row"],
                "end": p["_row"],
            }
        )
        starts_by_sheet[sheet].sort(key=lambda x: x["start"])

    image_by_name = {}
    for rec in image_records:
        if rec["kind"] != "product":
            continue
        blocks = starts_by_sheet.get(rec["sheet"], [])
        if not blocks:
            continue
        chosen = None
        for b in blocks:
            if b["start"] - 1 <= rec["row"] <= b["end"] + 1:
                chosen = b
                break
        if not chosen:
            priors = [b for b in blocks if b["start"] <= rec["row"] + 2]
            if priors:
                chosen = priors[-1]
        if not chosen:
            continue
        name = chosen["name"]
        prev = image_by_name.get(name)
        if not prev or score_image(rec) > score_image(prev):
            image_by_name[name] = rec

    for raw, meta in line_meta.items():
        name = display_name(raw)
        if name in image_by_name:
            continue
        candidates = [
            r
            for r in image_records
            if r["kind"] == "product"
            and r["sheet"] == meta["sheet"]
            and abs(r["row"] - meta["start_row"]) <= 3
        ]
        if candidates:
            image_by_name[name] = max(candidates, key=score_image)

    processed_meta = []
    for name, rec in sorted(image_by_name.items()):
        dest = PROC / f"{slugify(name)}.jpg"
        recolor_and_crop(Path(rec["path"]), dest)
        processed_meta.append(
            {
                "name": name,
                "file": dest.name,
                "path": str(dest),
                "source": rec["file"],
                "sheet": rec["sheet"],
                "row": rec["row"],
            }
        )
        print(f"processed: {name} <- {rec['file']}")

    blends = sorted({p["_blend"] for p in catalog})
    names = sorted({p["name"] for p in catalog})
    missing_img = [n for n in blends if n not in image_by_name]
    print(f"\nCatalog vitolas: {len(catalog)}")
    print(f"Distinct product names: {len(names)}")
    print(f"Distinct blends: {len(blends)}")
    print(f"Blends with images: {len(image_by_name)}")
    print(f"Missing images ({len(missing_img)}): {missing_img}")

    out_catalog = []
    for p in sorted(
        catalog, key=lambda x: (x["line"], x["name"], x.get("size_name") or "", x["length"])
    ):
        out_catalog.append(
            {
                "brand": p["brand"],
                "name": p["name"],
                "line": p["line"],
                "description": p["description"],
                "wrapper": p["wrapper"],
                "binder": p["binder"],
                "filler": p["filler"],
                "length": p["length"],
                "size_name": p.get("size_name") or "",
                "image": "",
            }
        )

    (ASSETS / "catalog.json").write_text(
        json.dumps(out_catalog, indent=2, ensure_ascii=False) + "\n"
    )
    (ROOT / "catalog.json").write_text(
        json.dumps(out_catalog, indent=2, ensure_ascii=False) + "\n"
    )
    (ROOT / "image-map.json").write_text(
        json.dumps(processed_meta, indent=2, ensure_ascii=False) + "\n"
    )
    (ROOT / "line-meta.json").write_text(
        json.dumps(line_meta, indent=2, ensure_ascii=False) + "\n"
    )
    singles = [
        {"name": m["name"], "blend": m["name"], "path": m["path"], "file": m["file"]}
        for m in processed_meta
    ]
    (ROOT / "singles.json").write_text(json.dumps(singles, indent=2) + "\n")

    blend_images = {m["name"]: m["path"] for m in processed_meta}
    (ROOT / "blend-images.json").write_text(
        json.dumps(blend_images, indent=2) + "\n"
    )

    print("\nWrote", ASSETS / "catalog.json")
    print("Sample:")
    for p in out_catalog[:6]:
        print(f"  {p['name']} | {p['length']} | {p['wrapper']}")


if __name__ == "__main__":
    main()
